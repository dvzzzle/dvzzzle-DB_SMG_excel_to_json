from openpyxl import load_workbook

wb = load_workbook(filename='E://Загрузки//Telegram Desktop//Текущая обработка//МФР_для_ДБ_от_18.03.xlsx', data_only=True)
ws = wb['6 - ОИВ КТ']

for row in ws.iter_rows(min_row=1, max_row=3, values_only=True):
    print(row)  # Отладочный вывод